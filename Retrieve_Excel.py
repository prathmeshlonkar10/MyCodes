import openpyxl
import sqlite3

con=sqlite3.connect('Excel.sqlite')
cur=con.cursor()

cur.executescript('''
DROP TABLE IF EXISTS examplexl;
CREATE TABLE examplexl (name TEXT, age INTEGER, designation TEXT, salary INTEGER)
''')
path='C:\\Users\\DELL\\Desktop\\exampleexcel.xlsx'
wb=openpyxl.load_workbook(path)
sheet=wb.active
maxc=sheet.max_column
maxr=sheet.max_row
print ('\nTotal number of rows:', maxr)
print ('Total number of columns:', maxc)
print()
for i in range (1, maxr+1):
    for j in range (1, maxc+1):
        cell=sheet.cell(row=i, column=j)
        val=cell.value
        print(val, end="\t")
        if j==1:
            name=val
        if j==2:
            age=val
        if j==3:
            des=val
        if j==4:
            sal=val
    cur.execute('''INSERT INTO examplexl (name, age, designation, salary) VALUES
                        (?,?,?,?)''', (name, age, des, sal))
    con.commit()
    print()
